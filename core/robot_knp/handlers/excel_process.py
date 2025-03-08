from pathlib import Path
import openpyxl
from openpyxl import Workbook


class ExcelManager:
    HEADERS = [
        'ЭЦП', 'Уведомления / Извещения / Решения', 'Сальдо лицевых счетов',
        'Запрос на получение сведений об отсутствии (наличии) задолженности',
        'Название организации'
    ]

    def __init__(self, file_path="data.xlsx"):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            self._create_file()

    def _create_file(self):
        """Создает новый файл Excel с заголовками"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(self.HEADERS)
        workbook.save(self.file_path)

    def _find_row(self, sheet, search_value):
        """Ищет строку по значению в первом столбце"""
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=1):
            if row[0].value == search_value:
                return row[0].row
        return None

    def write_data(self, data):
        """Записывает или обновляет данные в Excel"""
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        row_index = self._find_row(sheet, data[0]) or sheet.max_row + 1

        for col, value in enumerate(data, start=1):
            sheet.cell(row=row_index, column=col, value=value)

        with self.file_path.open("wb") as f:
            workbook.save(f)

    def update_data(self, folder_name, status, notification_status, user_info):
        """Обновляет данные в Excel"""
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        row_index = self._find_row(sheet, folder_name)
        if row_index is not None:
            sheet.cell(row=row_index, column=2, value=status)
            sheet.cell(row=row_index, column=3, value=notification_status)
            sheet.cell(row=row_index, column=5, value=user_info)
            with self.file_path.open("wb") as f:
                workbook.save(f)
        else:
            print(f"[ERROR] Не найдено: '{folder_name}' для обновления.")

# === --- ===
# if __name__ == "__main__":
#     excel = ExcelManager("report.xlsx")
#     excel.write_data(["123456", "Есть", "1000", "Запрос отправлен", "Компания X"])
#     excel.write_data(["123456", "Отсутствует", "2000", "Запрос отправлен", "Компания X"])
#     excel.write_data(["123233234456", "Недоимка", "2000", "Запрос отправлен", "Компания Xdd"])